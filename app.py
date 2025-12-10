#funcionamiento general
from flask import Flask, render_template, request, redirect, url_for, session, flash
from flask_mysqldb import MySQL
from config import Config
import MySQLdb.cursors
from datetime import datetime, date

#para el excel exportable:
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import send_file
from io import BytesIO

#=====================================================
# CONFIG GENERAL
#=====================================================
app = Flask(__name__)
app.config.from_object(Config)
app.secret_key = "STEELMAN_SUPER_KEY_2025"
mysql = MySQL(app)


#=====================================================
# CONTEXT PROCESSOR (Año dinámico)
#=====================================================
@app.context_processor
def inject_current_year():
    return {'current_year': datetime.now().year}


#=====================================================
# PÁGINAS PÚBLICAS
#=====================================================
@app.route('/')
def index():
    return render_template('publico/index.html')


@app.route('/rutas')
def rutas():
    return render_template('publico/rutas.html')


@app.route("/ruta/<string:codigo>")
def detalle_ruta(codigo):
    # Coordenadas de paradas por ruta (ejemplo)
    paraderos_data = {
        'A': [
            {'nombre': 'Plaza de Armas', 'lat': -16.4090, 'lng': -71.5375},
            {'nombre': 'Yanahuara', 'lat': -16.4120, 'lng': -71.5380},
            {'nombre': 'Cayma', 'lat': -16.4180, 'lng': -71.5320}
        ],
        'B': [
            {'nombre': 'Paucarpata', 'lat': -16.4005, 'lng': -71.5300},
            {'nombre': 'Av. Kennedy', 'lat': -16.4020, 'lng': -71.5280},
            {'nombre': 'Palacio Metropolitano', 'lat': -16.4080, 'lng': -71.5260}
        ],
        'C': [
            {'nombre': 'Umacollo', 'lat': -16.3960, 'lng': -71.5360},
            {'nombre': 'UNSA Ingenierías', 'lat': -16.3980, 'lng': -71.5350},
            {'nombre': 'Terminal Terrestre', 'lat': -16.4050, 'lng': -71.5330}
        ],
        'D': [
            {'nombre': 'Cerro Colorado', 'lat': -16.3900, 'lng': -71.5400},
            {'nombre': 'Zamacola', 'lat': -16.3920, 'lng': -71.5390},
            {'nombre': 'Av. La Marina', 'lat': -16.3980, 'lng': -71.5380}
        ],
        'E': [
            {'nombre': 'Sabandía', 'lat': -16.4080, 'lng': -71.5230},
            {'nombre': 'Characato', 'lat': -16.4070, 'lng': -71.5230},
            {'nombre': 'Cercado', 'lat': -16.4030, 'lng': -71.5210}
        ],
        'F': [
            {'nombre': 'Tiabaya', 'lat': -16.4120, 'lng': -71.5300},
            {'nombre': 'Hunter', 'lat': -16.4120, 'lng': -71.5290},
            {'nombre': 'Cercado', 'lat': -16.4090, 'lng': -71.5270}
        ]
    }
    
    paraderos = paraderos_data.get(codigo, [])
    return render_template("publico/ruta_detalle.html", codigo=codigo, paraderos_coords=paraderos)


@app.route('/contacto', methods=['GET','POST'])
def contacto():
    if request.method == 'POST':
        return render_template('publico/contacto.html',
                               enviado=True,
                               nombre=request.form.get('nombre'))
    return render_template('publico/contacto.html')

#=====================================================
# LOGIN / AUTENTICACIÓN 
#=====================================================
@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        usuario = request.form.get('usuario')
        clave = request.form.get('password')

        cursor = mysql.connection.cursor()
        cursor.execute("""
            SELECT id_empleado, tipo_supervisor, usuario
            FROM supervisor
            WHERE usuario=%s AND contraseña=%s
        """, (usuario, clave))

        cuenta = cursor.fetchone() 
        cursor.close()

        print("DEBUG consulta:", cuenta)

        if cuenta:
            session['id'] = cuenta['id_empleado'] 
            session['rol'] = cuenta['tipo_supervisor']
            session['usuario'] = cuenta['usuario']

            flash("✅ Bienvenido al sistema", "success")
            return redirect(url_for('panel'))

        flash("❌ Usuario o contraseña incorrectos", "danger")

    return render_template('autenticacion/login.html')




#=====================================================
# PANEL PRIVADO - DASHBOARD 
#=====================================================

@app.route('/panel')
def panel():
    if 'usuario' not in session:
        return redirect(url_for('login'))

    id_supervisor = session['id']
    
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    
    cursor.execute("""
        SELECT id_ruta 
        FROM supervisor_ruta
        WHERE id_empleado = %s AND fecha_fin IS NULL
    """, (id_supervisor,))
    ruta_data = cursor.fetchone()
    id_ruta = ruta_data['id_ruta'] if ruta_data else None

    
    kpi_buses = 0
    kpi_incidencias = 0
    kpi_recaudacion = "S/. 0.00"
    mensaje_ruta = ""
    nombre_ruta = "N/A"

    if id_ruta:
        
        cursor.execute("""
            SELECT COUNT(b.id_bus) AS total_buses
            FROM bus b
            JOIN bus_ruta br ON b.id_bus = br.id_bus
            WHERE br.id_ruta = %s AND br.fecha_desasignacion IS NULL
        """, (id_ruta,))
        kpi_buses = cursor.fetchone()['total_buses']
         
        cursor.execute("""
            SELECT SUM(monto_recaudado) AS recaudacion
            FROM caja
            WHERE id_ruta = %s 
              AND DATE(fecha) = CURDATE()
        """, (id_ruta,))
        
        recaudacion_hoy = cursor.fetchone()['recaudacion']
        
        if recaudacion_hoy is not None:
             kpi_recaudacion = f"S/. {recaudacion_hoy:,.2f}" 
        
    
        cursor.execute("SELECT letra FROM ruta WHERE id_ruta = %s", (id_ruta,))
        
        ruta_info = cursor.fetchone()
        
       
        nombre_ruta = ruta_info['letra'] if ruta_info and 'letra' in ruta_info else 'Ruta Desconocida'
        
        mensaje_ruta = f"✅ Administrando la Ruta: {nombre_ruta} (ID: {id_ruta})"
        
    else:
        mensaje_ruta = "🚫 No tienes una ruta activa asignada. Los KPIs de Ruta se muestran en 0."


    cursor.execute("""
        SELECT COUNT(id_incidencia) AS total_incidencias
        FROM incidencia
        WHERE estado = 'ABIERTA' OR estado = 'PENDIENTE'
    """)
    kpi_incidencias = cursor.fetchone()['total_incidencias']
        
    cursor.close()

    return render_template(
        'privado/dashboard.html',
        nombre_supervisor=session['usuario'],
        tipo_supervisor=session['rol'],
        kpi_buses=kpi_buses,
        kpi_incidencias=kpi_incidencias,
        kpi_recaudacion=kpi_recaudacion,
        mensaje_ruta=mensaje_ruta
    )



#=====================================================
# MODULOS DE BUSESS
#=====================================================

@app.route("/panel/buses", methods=["GET", "POST"])
def panel_buses():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    
    id_supervisor_logueado = session['id']
    cursor = mysql.connection.cursor()
    

    cursor.execute("SELECT tipo_supervisor FROM supervisor WHERE id_empleado = %s", (id_supervisor_logueado,))
    rol_data = cursor.fetchone()
    rol_supervisor = rol_data['tipo_supervisor'] if rol_data else None
    

    buses = []
    modelos = []
    id_ruta_asignada = None
    bus_a_editar = None 


    cursor.execute("""
        SELECT sr.id_ruta 
        FROM supervisor_ruta sr
        WHERE sr.id_empleado = %s AND sr.fecha_fin IS NULL
    """, (id_supervisor_logueado,))
    
    resultado_ruta = cursor.fetchone() 
    
    if not resultado_ruta:
        flash("🚫 No tienes rutas activas asignadas. No se muestran buses ni se permite el registro/edición.", "warning")
    else:
        id_ruta_asignada = resultado_ruta['id_ruta']


    if request.method == "POST":
        placa = request.form.get("placa")
        id_modelo_bus = request.form.get("id_modelo_bus")
        anio = request.form.get("anio") 
        revision = request.form.get("ultima_revision") 
        id_bus_editado = request.form.get("id_bus_editado") 

        if id_bus_editado:
            
            
            cursor.execute("""
                SELECT b.id_bus
                FROM bus b
                JOIN bus_ruta br ON b.id_bus = br.id_bus
                JOIN supervisor_ruta sr ON br.id_ruta = sr.id_ruta
                WHERE b.id_bus = %s 
                  AND sr.id_empleado = %s 
                  AND sr.fecha_fin IS NULL 
                  AND br.fecha_desasignacion IS NULL
            """, (id_bus_editado, id_supervisor_logueado))
            
            if not cursor.fetchone():
                flash("❌ Error de seguridad: Intento de editar un bus no asignado a tu ruta.", "danger")
            else:
                try:
                    
                    update_query = """
                        UPDATE bus 
                        SET placa = %s, 
                            año_fabricacion = %s, 
                            id_modelo_bus = %s
                    """
                    update_params = [placa, anio, id_modelo_bus]

                   
                    if rol_supervisor == 'General':
                        update_query += ", ultima_revision = %s"
                        update_params.append(revision)
                    
            
                    update_query += " WHERE id_bus = %s"
                    update_params.append(id_bus_editado)
                    
                    cursor.execute(update_query, update_params)
                    
                    mysql.connection.commit()
                    flash("✔️ Bus editado correctamente.", "success")
                except Exception as e:
                    mysql.connection.rollback()
                    flash(f"❌ Error al editar el bus: {str(e)}", "danger")

        else:
        
            if not id_ruta_asignada:
                flash("❌ No se puede registrar un bus sin tener una ruta asignada", "danger")
            else:
                cursor.execute("SELECT placa FROM bus WHERE placa=%s", (placa,))
                existe = cursor.fetchone()
                
                if existe:
                    flash("❌ Esa placa ya está registrada", "danger")
                else:
                    try:
                        
                        cursor.execute("""
                            INSERT INTO bus (placa, año_fabricacion, id_modelo_bus, id_almacen)
                            VALUES (%s, %s, %s, 1)
                        """, (placa, anio, id_modelo_bus))
                        
                       
                        cursor.execute("""
                            INSERT INTO bus_ruta (id_bus, id_ruta, fecha_asignacion)
                            VALUES (LAST_INSERT_ID(), %s, CURDATE())
                        """, (id_ruta_asignada,)) 
                        
                        mysql.connection.commit()
                        flash("✔️ Bus registrado y asignado a tu ruta correctamente", "success")
                    except Exception as e:
                        mysql.connection.rollback()
                        flash(f"❌ Error al registrar el bus: {str(e)}", "danger")

    if id_ruta_asignada:
        
        cursor.execute("""
            SELECT 
                mb.id_modelo_bus AS id, 
                mb.nombre AS modelo_nombre,
                ma.nombre AS marca_nombre
            FROM modelo_bus mb
            JOIN marca_bus ma ON mb.id_marca_bus = ma.id_marca_bus
            ORDER BY ma.nombre, mb.nombre
        """)
        modelos = cursor.fetchall()

        cursor.execute("""
            SELECT 
                b.id_bus, 
                b.placa,
                m.nombre AS modelo,
                ma.nombre AS marca,
                b.año_fabricacion AS año,
                b.ultima_revision AS revision
            FROM bus b
            JOIN modelo_bus m ON b.id_modelo_bus = m.id_modelo_bus
            JOIN marca_bus ma ON m.id_marca_bus = ma.id_marca_bus
            JOIN bus_ruta br ON b.id_bus = br.id_bus   
            WHERE br.id_ruta = %s                    
              AND br.fecha_desasignacion IS NULL     
        """, (id_ruta_asignada,))
        buses = cursor.fetchall()
        

        id_bus_editar = request.args.get('id_editar', type=int)

        if id_bus_editar:
   
            cursor.execute("""
                SELECT b.id_bus, b.placa, b.año_fabricacion, b.id_modelo_bus, b.ultima_revision
                FROM bus b
                JOIN bus_ruta br ON b.id_bus = br.id_bus
                JOIN supervisor_ruta sr ON br.id_ruta = sr.id_ruta
                WHERE b.id_bus = %s 
                  AND sr.id_empleado = %s 
                  AND sr.fecha_fin IS NULL 
                  AND br.fecha_desasignacion IS NULL
            """, (id_bus_editar, id_supervisor_logueado))
            
            bus_a_editar_data = cursor.fetchone()
            
            if bus_a_editar_data:
              
                bus_a_editar = bus_a_editar_data
            else:
                 flash("❌ El bus solicitado no existe o no está en tu ruta activa.", "danger")


    cursor.close()

 
    return render_template("privado/buses.html", 
                           buses=buses, 
                           modelos=modelos, 
                           bus_a_editar=bus_a_editar,
                           rol_supervisor=rol_supervisor)


###############################################
############# ELIMINAR BOTON BUSES#############
###############################################

@app.route("/panel/buses/eliminar/<int:id_bus>")
def panel_buses_eliminar(id_bus):
    if 'usuario' not in session:
        return redirect(url_for('login'))
    

    id_supervisor_logueado = session['id']
    cursor = mysql.connection.cursor()

   
    cursor.execute("""
        SELECT b.id_bus
        FROM bus b
        JOIN bus_ruta br ON b.id_bus = br.id_bus
        JOIN supervisor_ruta sr ON br.id_ruta = sr.id_ruta
        WHERE b.id_bus = %s 
          AND sr.id_empleado = %s 
          AND sr.fecha_fin IS NULL 
          AND br.fecha_desasignacion IS NULL
    """, (id_bus, id_supervisor_logueado))
    
    bus_valido = cursor.fetchone()

    if not bus_valido:
        flash("❌ Acceso denegado. El bus no está asignado a tu ruta activa.", "danger")
    else:
        try:
            
            cursor.execute("""
                UPDATE bus_ruta
                SET fecha_desasignacion = CURDATE()
                WHERE id_bus = %s AND fecha_desasignacion IS NULL
            """, (id_bus,))
            
            mysql.connection.commit()
            flash("✔️ Bus desasignado de tu ruta correctamente.", "success")
        except Exception as e:
            mysql.connection.rollback()
            flash(f"❌ Error al intentar desasignar el bus. {str(e)}", "danger")

    cursor.close()
    return redirect(url_for('panel_buses'))




#==============================================================================
# MODULO DEL PERSONAL
#==============================================================================

@app.route("/panel/personal", methods=['GET'])
def panel_personal():
    if 'usuario' not in session:
        return redirect(url_for('login'))

    return render_template("privado/personal.html")


######################################################
############# INTERFAZ PARA LOS CHOFERES #############
######################################################

@app.route("/panel/personal/choferes", methods=['GET'])
def panel_personal_choferes():
    if 'usuario' not in session:
        return redirect(url_for('login'))

    cursor = mysql.connection.cursor()
    id_supervisor = session['id']

 
    cursor.execute("""
        SELECT id_ruta
        FROM supervisor_ruta
        WHERE id_empleado=%s AND fecha_fin IS NULL
    """, (id_supervisor,))
    row = cursor.fetchone()

    if not row:
        cursor.close()
        return render_template("privado/personal_choferes.html",
                               choferes_existentes=[], id_ruta=None,
                               tipos_licencia=[], editar_personal=None)

    id_ruta = row['id_ruta']

  
    cursor.execute("""
        (
            SELECT e.id_empleado, p.nombre, p.apellido, p.dni,
                   ch.nro_licencia, tl.categoria, ch.años_experiencia
            FROM empleado e
            JOIN persona p ON e.id_persona=p.id_persona
            JOIN chofer ch ON e.id_empleado=ch.id_empleado
            JOIN tipo_licencia tl ON tl.id_tipo_licencia=ch.id_tipo_licencia
            WHERE e.id_empleado IN (
                SELECT DISTINCT ab.id_empleado
                FROM asignacion_bus ab
                JOIN bus b ON ab.id_bus=b.id_bus
                JOIN bus_ruta br ON b.id_bus=br.id_bus
                WHERE br.id_ruta=%s
            )
        )
        UNION
        (
            SELECT e.id_empleado, p.nombre, p.apellido, p.dni,
                   ch.nro_licencia, tl.categoria, ch.años_experiencia
            FROM empleado e
            JOIN persona p ON e.id_persona=p.id_persona
            JOIN chofer ch ON e.id_empleado=ch.id_empleado
            JOIN tipo_licencia tl ON tl.id_tipo_licencia=ch.id_tipo_licencia
            WHERE e.id_empleado NOT IN (SELECT id_empleado FROM asignacion_bus)
        )
    """, (id_ruta,))

    choferes = cursor.fetchall()

    cursor.execute("SELECT * FROM tipo_licencia ORDER BY categoria")
    tipos_licencia = cursor.fetchall()

    editar_personal = None

  
    if "editar" in request.args:
        cursor.execute("""
            SELECT e.id_empleado, p.nombre, p.apellido, p.dni, p.telefono,
                   e.sueldo,
                   ch.nro_licencia, ch.años_experiencia, ch.id_tipo_licencia,
                   ch.historial_infracciones
            FROM empleado e
            JOIN persona p ON e.id_persona=p.id_persona
            JOIN chofer ch ON ch.id_empleado=e.id_empleado
            WHERE e.id_empleado=%s
        """, (request.args["editar"],))
        editar_personal = cursor.fetchone()

   
    if "eliminar" in request.args:
        try:
            mysql.connection.begin()
            id_empleado = request.args["eliminar"]

            cursor.execute("DELETE FROM chofer WHERE id_empleado=%s", (id_empleado,))
            cursor.execute("DELETE FROM empleado WHERE id_empleado=%s", (id_empleado,))

            mysql.connection.commit()
            flash("✔ Registro eliminado.", "success")

        except Exception as e:
            mysql.connection.rollback()
            flash(f"❌ Error: {str(e)}", "danger")

        cursor.close()
        return redirect(url_for("panel_personal_choferes"))

    cursor.close()

    return render_template("privado/personal_choferes.html",
                           choferes_existentes=choferes,
                           id_ruta=id_ruta,
                           tipos_licencia=tipos_licencia,
                           editar_personal=editar_personal)


######################################################
############# REGISTRAR CHOFERES #####################
######################################################

@app.route("/registrar_personal_chofer", methods=['POST'])
def registrar_personal_chofer():
    cursor = mysql.connection.cursor()

    try:
        mysql.connection.begin()

        
        cursor.execute("""
            INSERT INTO persona(nombre, apellido, dni, telefono)
            VALUES (%s, %s, %s, %s)
        """, (
            request.form["nombre"],
            request.form["apellido"],
            request.form["dni"],
            request.form["telefono"]
        ))
        id_persona = cursor.lastrowid

        
        cursor.execute("""
            INSERT INTO empleado(id_persona, sueldo, fecha_ingreso)
            VALUES (%s, %s, %s)
        """, (
            id_persona,
            request.form["sueldo"],
            request.form["fecha_ingreso"]
        ))
        id_empleado = cursor.lastrowid

        
        cursor.execute("""
            INSERT INTO chofer(id_empleado, nro_licencia, años_experiencia,
                               id_tipo_licencia, historial_infracciones)
            VALUES (%s, %s, %s, %s, %s)
        """, (
            id_empleado,
            request.form["nro_licencia"],
            request.form["anos_experiencia"],
            request.form["id_tipo_licencia"],
            request.form["historial_infracciones"]
        ))

        mysql.connection.commit()
        flash("✔ Chofer registrado correctamente.", "success")

    except Exception as e:
        mysql.connection.rollback()
        flash(f"❌ Error: {str(e)}", "danger")

    cursor.close()
    return redirect(url_for("panel_personal_choferes"))



######################################################
############# ACTUALIZAR CHOFERES ####################
######################################################

@app.route("/actualizar_personal_chofer", methods=['POST'])
def actualizar_personal_chofer():
    cursor = mysql.connection.cursor()
    try:
        mysql.connection.begin()
        id_empleado = request.form["id_empleado"]

        
        cursor.execute("""
            UPDATE persona p
            JOIN empleado e ON p.id_persona=e.id_persona
            SET p.nombre=%s, p.apellido=%s, p.dni=%s, p.telefono=%s
            WHERE e.id_empleado=%s
        """, (
            request.form["nombre"],
            request.form["apellido"],
            request.form["dni"],
            request.form["telefono"],
            id_empleado
        ))

        
        cursor.execute("""
            UPDATE empleado
            SET sueldo=%s
            WHERE id_empleado=%s
        """, (request.form["sueldo"], id_empleado))

        
        cursor.execute("""
            UPDATE chofer
            SET nro_licencia=%s,
                años_experiencia=%s,
                id_tipo_licencia=%s,
                historial_infracciones=%s
            WHERE id_empleado=%s
        """, (
            request.form["nro_licencia"],
            request.form["anos_experiencia"],
            request.form["id_tipo_licencia"],
            request.form["historial_infracciones"],
            id_empleado
        ))

        mysql.connection.commit()
        flash("✔ Datos del chofer actualizados.", "success")
    except Exception as e:
        mysql.connection.rollback()
        flash(f"❌ Error: {str(e)}", "danger")
    finally:
        cursor.close()

    return redirect(url_for("panel_personal_choferes"))




######################################################
############# INTERFAZ PARA COBRADORES ###############
######################################################

@app.route("/panel/personal/cobradores", methods=['GET'])
def panel_personal_cobradores():
    if 'usuario' not in session:
        return redirect(url_for('login'))

    cursor = mysql.connection.cursor()
    id_supervisor = session['id']

   
    cursor.execute("""
        SELECT id_ruta
        FROM supervisor_ruta
        WHERE id_empleado=%s AND fecha_fin IS NULL
    """, (id_supervisor,))
    row = cursor.fetchone()

    if not row:
        cursor.close()
        return render_template("privado/personal_cobradores.html",
                               cobradores_existentes=[], id_ruta=None,
                               editar_personal=None,
                               lista_idiomas=[],
                               idiomas_asociados=[])

    id_ruta = row['id_ruta']

    
    cursor.execute("""
        (
            SELECT e.id_empleado, p.nombre, p.apellido, p.dni, p.telefono, e.sueldo
            FROM empleado e
            JOIN persona p ON e.id_persona=p.id_persona
            JOIN cobrador c ON c.id_empleado=e.id_empleado
            WHERE e.id_empleado IN (
                SELECT DISTINCT ab.id_empleado
                FROM asignacion_bus ab
                JOIN bus b ON ab.id_bus=b.id_bus
                JOIN bus_ruta br ON b.id_bus=br.id_bus
                WHERE br.id_ruta=%s
            )
        )
        UNION
        (
            SELECT e.id_empleado, p.nombre, p.apellido, p.dni, p.telefono, e.sueldo
            FROM empleado e
            JOIN persona p ON e.id_persona=p.id_persona
            JOIN cobrador c ON c.id_empleado=e.id_empleado
            WHERE e.id_empleado NOT IN (SELECT id_empleado FROM asignacion_bus)
        )
    """, (id_ruta,))
    
    cobradores = cursor.fetchall()

    
    cursor.execute("SELECT id_idioma, nombre FROM idioma")
    lista_idiomas = cursor.fetchall()

    editar_personal = None
    idiomas_asociados = []

    
    if "editar" in request.args:
        cursor.execute("""
            SELECT e.id_empleado, p.nombre, p.apellido, p.dni, p.telefono,
                   e.sueldo
            FROM empleado e
            JOIN persona p ON e.id_persona=p.id_persona
            JOIN cobrador c ON c.id_empleado=e.id_empleado
            WHERE e.id_empleado=%s
        """, (request.args["editar"],))
        editar_personal = cursor.fetchone()

        cursor.execute("""
            SELECT id_idioma
            FROM cobrador_idioma
            WHERE id_empleado=%s
        """, (editar_personal["id_empleado"],))
        idiomas_asociados = [row["id_idioma"] for row in cursor.fetchall()]

    
    if "eliminar" in request.args:
        try:
            mysql.connection.begin()
            id_empleado = request.args["eliminar"]

            cursor.execute("DELETE FROM cobrador_idioma WHERE id_empleado=%s", (id_empleado,))
            cursor.execute("DELETE FROM cobrador WHERE id_empleado=%s", (id_empleado,))
            cursor.execute("DELETE FROM empleado WHERE id_empleado=%s", (id_empleado,))

            mysql.connection.commit()
            flash("✔ Registro eliminado.", "success")

        except Exception as e:
            mysql.connection.rollback()
            flash(f"❌ Error: {str(e)}", "danger")

        cursor.close()
        return redirect(url_for("panel_personal_cobradores"))

    cursor.close()

    
    for c in cobradores:
        cursor2 = mysql.connection.cursor()
        cursor2.execute("""
            SELECT i.nombre
            FROM cobrador_idioma ci
            JOIN idioma i ON ci.id_idioma=i.id_idioma
            WHERE ci.id_empleado=%s
        """, (c["id_empleado"],))
        idiomas = [row["nombre"] for row in cursor2.fetchall()]
        c["idiomas"] = ", ".join(idiomas)
        cursor2.close()

    return render_template("privado/personal_cobradores.html",
                           cobradores_existentes=cobradores,
                           id_ruta=id_ruta,
                           editar_personal=editar_personal,
                           lista_idiomas=lista_idiomas,
                           idiomas_asociados=idiomas_asociados)


######################################################
############# REGISTRAR COBRADORES ###################
######################################################

@app.route("/registrar_personal_cobrador", methods=['POST'])
def registrar_personal_cobrador():
    cursor = mysql.connection.cursor()

    try:
        mysql.connection.begin()

        cursor.execute("""
            INSERT INTO persona(nombre, apellido, dni, telefono)
            VALUES (%s, %s, %s, %s)
        """, (
            request.form["nombre"],
            request.form["apellido"],
            request.form["dni"],
            request.form["telefono"]
        ))
        id_persona = cursor.lastrowid

        cursor.execute("""
            INSERT INTO empleado(id_persona, sueldo, fecha_ingreso, tipo_empleado)
            VALUES (%s, %s, %s, 'COBRADOR')
        """, (
            id_persona,
            request.form["sueldo"],
            request.form["fecha_ingreso"]
        ))
        id_empleado = cursor.lastrowid

      
        cursor.execute("""
            INSERT INTO cobrador(id_empleado)
            VALUES (%s)
        """, (id_empleado,))

        
        idiomas = request.form.getlist("idiomas")
        for id_idioma in idiomas:
            cursor.execute("""
                INSERT INTO cobrador_idioma(id_empleado, id_idioma)
                VALUES (%s, %s)
            """, (id_empleado, id_idioma))

        mysql.connection.commit()
        flash("✔ Cobrador registrado correctamente.", "success")

    except Exception as e:
        mysql.connection.rollback()
        flash(f"❌ Error: {str(e)}", "danger")

    cursor.close()
    return redirect(url_for("panel_personal_cobradores"))


######################################################
############# ACTUALIZAR COBRADORES ##################
######################################################

@app.route("/actualizar_personal_cobrador", methods=['POST'])
def actualizar_personal_cobrador():
    cursor = mysql.connection.cursor()

    try:
        mysql.connection.begin()

        id_empleado = request.form["id_empleado"]

        cursor.execute("""
            UPDATE persona p
            JOIN empleado e ON p.id_persona=e.id_persona
            SET p.nombre=%s, p.apellido=%s, p.dni=%s, p.telefono=%s
            WHERE e.id_empleado=%s
        """, (
            request.form["nombre"],
            request.form["apellido"],
            request.form["dni"],
            request.form["telefono"],
            id_empleado
        ))

        cursor.execute("""
            UPDATE empleado
            SET sueldo=%s
            WHERE id_empleado=%s
        """, (request.form["sueldo"], id_empleado))

        
        cursor.execute("DELETE FROM cobrador_idioma WHERE id_empleado=%s", (id_empleado,))
        idiomas = request.form.getlist("idiomas")
        for id_idioma in idiomas:
            cursor.execute("""
                INSERT INTO cobrador_idioma(id_empleado, id_idioma)
                VALUES (%s, %s)
            """, (id_empleado, id_idioma))

        mysql.connection.commit()
        flash("✔ Datos de cobrador actualizados.", "success")

    except Exception as e:
        mysql.connection.rollback()
        flash(f"❌ Error: {str(e)}", "danger")

    cursor.close()
    return redirect(url_for("panel_personal_cobradores"))





#=============================================
# MODULO DE INCIDENCIAS
#=============================================

@app.route("/panel/incidencias", methods=['GET'])
def panel_incidencias():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    return render_template("privado/incidencias.html")


######################################################
############# INCIDENCIAS DISCIPLINARIAS #############
######################################################

@app.route("/panel/incidencias/disciplinarias", methods=['GET'])
def panel_incidencias_disciplinarias():
    if 'usuario' not in session:
        return redirect(url_for('login'))

    cursor = mysql.connection.cursor()

    id_supervisor = session['id']

    cursor.execute("""
        SELECT id_ruta
        FROM supervisor_ruta
        WHERE id_empleado=%s AND fecha_fin IS NULL
    """, (id_supervisor,))
    row = cursor.fetchone()
    id_ruta = row['id_ruta'] if row else None

    # ========== NUEVO: obtener buses asignados a la ruta mediante bus_ruta ==========
    buses = []
    if id_ruta:
        cursor.execute("""
            SELECT b.id_bus, b.placa
            FROM bus b
            JOIN bus_ruta br ON b.id_bus = br.id_bus
            WHERE br.id_ruta = %s
              AND (br.fecha_desasignacion IS NULL OR br.fecha_desasignacion >= CURDATE())
        """, (id_ruta,))
        buses = cursor.fetchall()

    # ========== listar incidencias ==========
    cursor.execute("""
        SELECT i.id_incidencia, i.fecha, i.descripcion, i.estado, i.id_bus,
               id.tipo_disciplinaria, id.sancion
        FROM incidencia i
        JOIN incidencia_disciplinaria id
        ON i.id_incidencia=id.id_incidencia
        ORDER BY i.fecha DESC
    """)

    incidencias = cursor.fetchall()

    # ========== edición ==========
    editar_id = request.args.get('editar')
    editar_incidencia = None

    if editar_id:
        cursor.execute("""
            SELECT i.id_incidencia, i.fecha, i.descripcion, i.estado, i.id_bus,
                   id.tipo_disciplinaria, id.sancion
            FROM incidencia i
            JOIN incidencia_disciplinaria id
            ON i.id_incidencia=id.id_incidencia
            WHERE i.id_incidencia=%s
        """, (editar_id,))
        editar_incidencia = cursor.fetchone()

        # Si el bus asociado a la incidencia no está en la lista de buses (p.ej. desasignado),
        # lo agregamos para que el select muestre el valor actual y no rompa el formulario.
        try:
            if editar_incidencia and editar_incidencia.get('id_bus'):
                editar_bus_id = editar_incidencia.get('id_bus')
                found = False
                for b in buses:
                    # soportar filas tipo dict o tupla
                    if isinstance(b, dict):
                        if b.get('id_bus') == editar_bus_id:
                            found = True
                            break
                    else:
                        # tuple/list: asumimos (id_bus, placa)
                        if len(b) >= 1 and b[0] == editar_bus_id:
                            found = True
                            break
                if not found:
                    cursor.execute("SELECT id_bus, placa FROM bus WHERE id_bus=%s", (editar_bus_id,))
                    extra = cursor.fetchone()
                    if extra:
                        buses.append(extra)
        except Exception:
            # no rompemos si algo raro sucede; simplemente no añadimos el extra
            pass

    cursor.close()

    return render_template(
        "privado/incidencias_disciplinarias.html",
        incidencias=incidencias,
        id_ruta=id_ruta,
        buses=buses,       # ← enviado al HTML
        hoy=date.today().isoformat(),
        editar_incidencia=editar_incidencia
    )


################################################################
############# REGISTRAR INCIDENCIAS DISCIPLINARIAS #############
################################################################

@app.route("/registrar_incidencia_disciplinaria", methods=['POST'])
def registrar_incidencia_disciplinaria():

    cursor = mysql.connection.cursor()

    try:
        try:
            fecha_obj = datetime.strptime(request.form["fecha"], "%Y-%m-%d").date()
        except ValueError:
            flash("❌ Fecha inválida. Usa formato YYYY-MM-DD.", "danger")
            return redirect(url_for("panel_incidencias_disciplinarias"))

        if fecha_obj < date(2000, 1, 1) or fecha_obj > date.today():
            flash("❌ Fecha fuera del rango permitido (2000 hasta hoy).", "danger")
            return redirect(url_for("panel_incidencias_disciplinarias"))

        mysql.connection.begin()

        cursor.execute("""
            INSERT INTO incidencia(fecha, descripcion, estado, id_bus)
            VALUES (%s, %s, %s, %s)
        """, (
            fecha_obj,
            request.form["descripcion"],
            request.form["estado"],
            request.form["id_bus"]
        ))

        id_incidencia = cursor.lastrowid

        cursor.execute("""
            INSERT INTO incidencia_disciplinaria(id_incidencia, tipo_disciplinaria, sancion)
            VALUES (%s, %s, %s)
        """, (
            id_incidencia,
            request.form["tipo_disciplinaria"],
            request.form["sancion"]
        ))

        mysql.connection.commit()
        flash("✔ Incidencia disciplinaria registrada.", "success")

    except Exception as e:
        mysql.connection.rollback()
        flash(f"❌ Error: {str(e)}", "danger")

    finally:
        cursor.close()

    return redirect(url_for("panel_incidencias_disciplinarias"))


#################################################################
############# ACTUALIZAR INCIDENCIAS DISCIPLINARIAS #############
#################################################################

@app.route("/actualizar_incidencia_disciplinaria", methods=['POST'])
def actualizar_incidencia_disciplinaria():

    cursor = mysql.connection.cursor()

    try:
        id_incidencia = request.form["id_incidencia"]

        try:
            fecha_obj = datetime.strptime(request.form["fecha"], "%Y-%m-%d").date()
        except ValueError:
            flash("❌ Fecha inválida. Usa formato YYYY-MM-DD.", "danger")
            return redirect(url_for("panel_incidencias_disciplinarias"))

        if fecha_obj < date(2000, 1, 1) or fecha_obj > date.today():
            flash("❌ Fecha fuera del rango permitido (2000 hasta hoy).", "danger")
            return redirect(url_for("panel_incidencias_disciplinarias"))

        mysql.connection.begin()

        cursor.execute("""
            UPDATE incidencia
            SET fecha=%s, descripcion=%s, estado=%s, id_bus=%s
            WHERE id_incidencia=%s
        """, (
            fecha_obj,
            request.form["descripcion"],
            request.form["estado"],
            request.form["id_bus"],
            id_incidencia
        ))

        cursor.execute("""
            UPDATE incidencia_disciplinaria
            SET tipo_disciplinaria=%s, sancion=%s
            WHERE id_incidencia=%s
        """, (
            request.form["tipo_disciplinaria"],
            request.form["sancion"],
            id_incidencia
        ))

        mysql.connection.commit()
        flash("✔ Incidencia disciplinaria actualizada.", "success")

    except Exception as e:
        mysql.connection.rollback()
        flash(f"❌ Error: {str(e)}", "danger")

    finally:
        cursor.close()

    return redirect(url_for("panel_incidencias_disciplinarias"))


###############################################################
############# ELIMINAR INCIDENCIAS DISCIPLINARIAS #############
###############################################################

@app.route("/eliminar_incidencia_disciplinaria/<int:id_incidencia>", methods=['GET'])
def eliminar_incidencia_disciplinaria(id_incidencia):

    cursor = mysql.connection.cursor()

    try:
        mysql.connection.begin()

        cursor.execute("DELETE FROM incidencia_disciplinaria WHERE id_incidencia=%s", (id_incidencia,))
        cursor.execute("DELETE FROM incidencia WHERE id_incidencia=%s", (id_incidencia,))

        mysql.connection.commit()
        flash("✔ Incidencia disciplinaria eliminada.", "success")

    except Exception as e:
        mysql.connection.rollback()
        flash(f"❌ Error: {str(e)}", "danger")

    finally:
        cursor.close()

    return redirect(url_for("panel_incidencias_disciplinarias"))






#=============================================
# INCIDENCIAS OPERATIVAS
#=============================================
@app.route("/panel/incidencias/operativas", methods=['GET'])
def panel_incidencias_operativas():
    if 'usuario' not in session:
        return redirect(url_for('login'))

    cursor = mysql.connection.cursor()
    id_supervisor = session['id']

    cursor.execute("""
        SELECT id_ruta
        FROM supervisor_ruta
        WHERE id_empleado=%s AND fecha_fin IS NULL
    """, (id_supervisor,))
    row = cursor.fetchone()
    id_ruta = row['id_ruta'] if row else None

    
    if id_ruta:
        cursor.execute("""
            SELECT i.id_incidencia, i.fecha, i.descripcion, i.estado,
                   io.gravedad, io.costo, io.requiere_seguro
            FROM incidencia i
            JOIN incidencia_operativa io ON i.id_incidencia=io.id_incidencia
            ORDER BY i.fecha DESC
        """)
    else:
        cursor.execute("""
            SELECT i.id_incidencia, i.fecha, i.descripcion, i.estado,
                   io.gravedad, io.costo, io.requiere_seguro
            FROM incidencia i
            JOIN incidencia_operativa io ON i.id_incidencia=io.id_incidencia
            ORDER BY i.fecha DESC
        """)
    incidencias = cursor.fetchall()

    
    editar_id = request.args.get('editar')
    editar_incidencia = None
    if editar_id:
        cursor.execute("""
            SELECT i.id_incidencia, i.fecha, i.descripcion, i.estado,
                   io.gravedad, io.costo, io.requiere_seguro
            FROM incidencia i
            JOIN incidencia_operativa io ON i.id_incidencia=io.id_incidencia
            WHERE i.id_incidencia=%s
        """, (editar_id,))
        editar_incidencia = cursor.fetchone()

    cursor.close()

   
    return render_template("privado/incidencias_operativas.html",
                           incidencias=incidencias,
                           id_ruta=id_ruta,
                           hoy=date.today().isoformat(),
                           editar_incidencia=editar_incidencia)




############################################################
############# REGISTRAR INCIDENCIAS OPERATIVAS #############
############################################################

@app.route("/registrar_incidencia_operativa", methods=['POST'])
def registrar_incidencia_operativa():
    cursor = mysql.connection.cursor()
    try:
        try:
            fecha_obj = datetime.strptime(request.form["fecha"], "%Y-%m-%d").date()
        except ValueError:
            flash("❌ Fecha inválida. Usa formato YYYY-MM-DD.", "danger")
            return redirect(url_for("panel_incidencias_operativas"))

        if fecha_obj < date(2000, 1, 1) or fecha_obj > date.today():
            flash("❌ Fecha fuera del rango permitido (2000 hasta hoy).", "danger")
            return redirect(url_for("panel_incidencias_operativas"))

        mysql.connection.begin()

        cursor.execute("""
            INSERT INTO incidencia(fecha, descripcion, estado)
            VALUES (%s, %s, %s)
        """, (
            fecha_obj,
            request.form["descripcion"],
            request.form["estado"]
        ))
        id_incidencia = cursor.lastrowid

        requiere_seguro = 1 if request.form.get("requiere_seguro") == "on" else 0
        costo = request.form.get("costo") or 0

        cursor.execute("""
            INSERT INTO incidencia_operativa(id_incidencia, gravedad, costo, requiere_seguro)
            VALUES (%s, %s, %s, %s)
        """, (
            id_incidencia,
            request.form["gravedad"],
            costo,
            requiere_seguro
        ))

        mysql.connection.commit()
        flash("✔ Incidencia operativa registrada.", "success")
    except Exception as e:
        mysql.connection.rollback()
        flash(f"❌ Error: {str(e)}", "danger")
    finally:
        cursor.close()

    return redirect(url_for("panel_incidencias_operativas"))



#############################################################
############# ACTUALIZAR INCIDENCIAS OPERATIVAS #############
#############################################################

@app.route("/actualizar_incidencia_operativa", methods=['POST'])
def actualizar_incidencia_operativa():
    cursor = mysql.connection.cursor()
    try:
        id_incidencia = request.form["id_incidencia"]

        try:
            fecha_obj = datetime.strptime(request.form["fecha"], "%Y-%m-%d").date()
        except ValueError:
            flash("❌ Fecha inválida. Usa formato YYYY-MM-DD.", "danger")
            return redirect(url_for("panel_incidencias_operativas"))

        if fecha_obj < date(2000, 1, 1) or fecha_obj > date.today():
            flash("❌ Fecha fuera del rango permitido (2000 hasta hoy).", "danger")
            return redirect(url_for("panel_incidencias_operativas"))

        mysql.connection.begin()

        cursor.execute("""
            UPDATE incidencia
            SET fecha=%s, descripcion=%s, estado=%s
            WHERE id_incidencia=%s
        """, (
            fecha_obj,
            request.form["descripcion"],
            request.form["estado"],
            id_incidencia
        ))

        requiere_seguro = 1 if request.form.get("requiere_seguro") == "on" else 0
        costo = request.form.get("costo") or 0

        cursor.execute("""
            UPDATE incidencia_operativa
            SET gravedad=%s, costo=%s, requiere_seguro=%s
            WHERE id_incidencia=%s
        """, (
            request.form["gravedad"],
            costo,
            requiere_seguro,
            id_incidencia
        ))

        mysql.connection.commit()
        flash("✔ Incidencia operativa actualizada.", "success")
    except Exception as e:
        mysql.connection.rollback()
        flash(f"❌ Error: {str(e)}", "danger")
    finally:
        cursor.close()

    return redirect(url_for("panel_incidencias_operativas"))





############################################################
############# ELIMINAR INCIDENCIAS OPERATIVAS ##############
############################################################

@app.route("/eliminar_incidencia_operativa/<int:id_incidencia>", methods=['GET'])
def eliminar_incidencia_operativa(id_incidencia):
    cursor = mysql.connection.cursor()
    try:
        mysql.connection.begin()
        cursor.execute("DELETE FROM incidencia_operativa WHERE id_incidencia=%s", (id_incidencia,))
        cursor.execute("DELETE FROM incidencia WHERE id_incidencia=%s", (id_incidencia,))
        mysql.connection.commit()
        flash("✔ Incidencia operativa eliminada.", "success")
    except Exception as e:
        mysql.connection.rollback()
        flash(f"❌ Error: {str(e)}", "danger")
    finally:
        cursor.close()

    return redirect(url_for("panel_incidencias_operativas"))







#=============================================
# PANEL DE CAJA
#=============================================

@app.route("/panel/caja", methods=['GET'])
def panel_caja():
    if 'usuario' not in session:
        return redirect(url_for('login'))

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor) 
    id_supervisor = session['id']

    cursor.execute("""
        SELECT id_ruta
        FROM supervisor_ruta
        WHERE id_empleado=%s AND fecha_fin IS NULL
    """, (id_supervisor,))
    row = cursor.fetchone()
    id_ruta = row['id_ruta'] if row else None

    cajas, empleados, buses = [], [], []

    if id_ruta:
       
        cursor.execute("""
            SELECT c.id_caja, c.fecha, c.monto_recaudado, c.observacion,
                   e.id_empleado, p.nombre, p.apellido, b.id_bus, b.placa
            FROM caja c
            JOIN empleado e ON c.id_empleado = e.id_empleado
            JOIN persona p ON e.id_persona = p.id_persona
            JOIN bus b ON c.id_bus = b.id_bus
            WHERE c.id_ruta=%s
            ORDER BY c.fecha DESC
        """, (id_ruta,))
        cajas = cursor.fetchall()


        cursor.execute("""
            (
                SELECT DISTINCT e.id_empleado, p.nombre, p.apellido
                FROM persona p
                JOIN empleado e ON e.id_persona = p.id_persona
                JOIN cobrador co ON e.id_empleado = co.id_empleado 
                WHERE e.id_empleado IN (
                    SELECT ab.id_empleado
                    FROM asignacion_bus ab
                    JOIN bus_ruta br ON ab.id_bus = br.id_bus
                    WHERE br.id_ruta = %s
                )
            )
            UNION
            (
                SELECT e.id_empleado, p.nombre, p.apellido
                FROM persona p
                JOIN empleado e ON e.id_persona = p.id_persona
                JOIN cobrador co ON e.id_empleado = co.id_empleado 
                WHERE e.id_empleado NOT IN (SELECT id_empleado FROM asignacion_bus)
            )
        """, (id_ruta,))
        empleados = cursor.fetchall()

       
        cursor.execute("""
            SELECT b.id_bus, b.placa
            FROM bus b
            JOIN bus_ruta br ON b.id_bus = br.id_bus
            WHERE br.id_ruta = %s
        """, (id_ruta,))
        buses = cursor.fetchall()

    cursor.close()
    hoy = datetime.today().strftime("%Y-%m-%d")

    return render_template("privado/caja.html",
                           cajas=cajas, 
                           empleados=empleados,
                           buses=buses,
                           id_ruta=id_ruta,
                           editar_caja=None,
                           hoy=hoy)




#################################################
############# REGISTRAR RECAUDACION #############
#################################################

@app.route("/registrar_caja", methods=['POST'])
def registrar_caja():
    if 'usuario' not in session:
        return redirect(url_for('login'))

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor) 
    id_supervisor = session['id']

    try:
        cursor.execute("""
            SELECT id_ruta
            FROM supervisor_ruta
            WHERE id_empleado=%s AND fecha_fin IS NULL
        """, (id_supervisor,))
        row = cursor.fetchone()
        id_ruta = row['id_ruta'] if row else None

        if not id_ruta:
            flash("❌ No tienes una ruta activa.", "danger")
            return redirect(url_for("panel_caja"))

        fecha = request.form.get("fecha")
        monto = request.form.get("monto")
        observacion = request.form.get("observacion")
        id_empleado = request.form.get("id_empleado")
        id_bus = request.form.get("id_bus")

        if not fecha or not monto or not id_empleado or not id_bus:
            flash("❌ Debes completar todos los campos obligatorios.", "danger")
            return redirect(url_for("panel_caja"))

        fecha_obj = datetime.strptime(fecha, "%Y-%m-%d").date()
        monto = float(monto)

        mysql.connection.begin()
        cursor.execute("""
            INSERT INTO caja(id_empleado, id_ruta, id_bus, fecha, monto_recaudado, observacion)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (id_empleado, id_ruta, id_bus, fecha_obj, monto, observacion))
        mysql.connection.commit()
        flash("✔ Recaudación registrada correctamente.", "success")

    except Exception as e:
        mysql.connection.rollback()
        flash(f"❌ Error: {str(e)}", "danger")
    finally:
        cursor.close()

    return redirect(url_for("panel_caja"))



#################################################
############# ELIMINAR RECAUDACION #############
#################################################

@app.route("/eliminar_caja/<int:id_caja>", methods=['GET'])
def eliminar_caja(id_caja):
    if 'usuario' not in session:
        return redirect(url_for('login'))

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor) 
    try:
        mysql.connection.begin()
        cursor.execute("DELETE FROM caja WHERE id_caja=%s", (id_caja,))
        mysql.connection.commit()
        flash("✔ Recaudación eliminada.", "success")
    except Exception as e:
        mysql.connection.rollback()
        flash(f"❌ Error: {str(e)}", "danger")
    finally:
        cursor.close()

    return redirect(url_for("panel_caja"))



#################################################
############# EDITAR RECAUDACION ################
#################################################

@app.route("/editar_caja/<int:id_caja>", methods=['GET'])
def editar_caja(id_caja):
    if 'usuario' not in session:
        return redirect(url_for('login'))

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor) 
    id_supervisor = session['id']

    cursor.execute("""
        SELECT id_ruta
        FROM supervisor_ruta
        WHERE id_empleado=%s AND fecha_fin IS NULL
    """, (id_supervisor,))
    row = cursor.fetchone()
    id_ruta = row['id_ruta'] if row else None

    
    cursor.execute("SELECT * FROM caja WHERE id_caja=%s", (id_caja,))
    editar_caja = cursor.fetchone()

    
    cajas = []
    if id_ruta:
        cursor.execute("""
            SELECT c.id_caja, c.fecha, c.monto_recaudado, c.observacion,
                   e.id_empleado, p.nombre, p.apellido, b.id_bus, b.placa
            FROM caja c
            JOIN empleado e ON c.id_empleado = e.id_empleado
            JOIN persona p ON e.id_persona = p.id_persona
            JOIN bus b ON c.id_bus = b.id_bus
            WHERE c.id_ruta=%s
            ORDER BY c.fecha DESC
        """, (id_ruta,))
        cajas = cursor.fetchall()
    
    
    cursor.execute("""
        (
            SELECT DISTINCT e.id_empleado, p.nombre, p.apellido
            FROM persona p
            JOIN empleado e ON e.id_persona = p.id_persona
            JOIN cobrador co ON e.id_empleado = co.id_empleado 
            WHERE e.id_empleado IN (
                SELECT ab.id_empleado
                FROM asignacion_bus ab
                JOIN bus_ruta br ON ab.id_bus = br.id_bus
                WHERE br.id_ruta = %s
            )
        )
        UNION
        (
            SELECT e.id_empleado, p.nombre, p.apellido
            FROM persona p
            JOIN empleado e ON e.id_persona = p.id_persona
            JOIN cobrador co ON e.id_empleado = co.id_empleado 
            WHERE e.id_empleado NOT IN (SELECT id_empleado FROM asignacion_bus)
        )
    """, (id_ruta,)) 
    empleados = cursor.fetchall()

    cursor.execute("""
        SELECT b.id_bus, b.placa
        FROM bus b
        JOIN bus_ruta br ON b.id_bus = br.id_bus
        WHERE br.id_ruta = %s
    """, (id_ruta,))
    buses = cursor.fetchall()

    cursor.close()
    hoy = datetime.today().strftime("%Y-%m-%d")

    return render_template("privado/caja.html",
                           cajas=cajas,
                           empleados=empleados,
                           buses=buses,
                           id_ruta=id_ruta,
                           editar_caja=editar_caja,
                           hoy=hoy)



##################################################
############# ACTUALIZAR RECAUDACION #############
##################################################

@app.route("/actualizar_caja", methods=['POST'])
def actualizar_caja():
    if 'usuario' not in session:
        return redirect(url_for('login'))

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor) 
    try:
        id_caja = request.form["id_caja"]
        fecha = request.form["fecha"]
        monto = float(request.form["monto"])
        observacion = request.form["observacion"]
        id_empleado = request.form["id_empleado"]
        id_bus = request.form["id_bus"]

        mysql.connection.begin()
        cursor.execute("""
            UPDATE caja
            SET fecha=%s, monto_recaudado=%s, observacion=%s,
                id_empleado=%s, id_bus=%s
            WHERE id_caja=%s
        """, (fecha, monto, observacion, id_empleado, id_bus, id_caja))
        mysql.connection.commit()
        flash("✔ Recaudación actualizada correctamente.", "success")

    except Exception as e:
        mysql.connection.rollback()
        flash(f"❌ Error al actualizar: {str(e)}", "danger")
    finally:
        cursor.close()

    return redirect(url_for("panel_caja"))


#=====================================================
# REPORTE EXTRAIBLE EXCEL
#=====================================================

@app.route("/panel/reportes")
def panel_reportes():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    return render_template("privado/reportes.html")


#=====================================================
# CERRAR SESION
#=====================================================
@app.route('/logout')
def logout():
    session.clear()
    flash("Sesión cerrada correctamente", "info")
    return redirect(url_for('login'))




#=====================================================
# EXPORTAR EXCEL
#=====================================================

@app.route('/exportar_excel')
def exportar_excel():
    cursor = mysql.connection.cursor()

    cursor.execute("SELECT DATABASE()")
    db_row = cursor.fetchone()
    db_name = db_row[list(db_row.keys())[0]] if db_row else "bd_stelman_buses"

    
    wb = Workbook()
    wb.remove(wb.active)

   
    title_font = Font(size=20, bold=True)
    subtitle_font = Font(size=12, italic=True)
    header_font = Font(bold=True)
    pk_fill = PatternFill("solid", fgColor="FFD966")        
    fk_fill = PatternFill("solid", fgColor="C9DAF8")      
    ddl_fill = PatternFill("solid", fgColor="FFF2CC")    
    index_header_fill = PatternFill("solid", fgColor="C6E0B4")
    alt_fill = PatternFill("solid", fgColor="F2F2F2")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_align = Alignment(horizontal="center", vertical="center")

    def auto_width(sheet):
        
        for col in sheet.columns:
            max_len = 0
            try:
                col_letter = get_column_letter(col[0].column)
            except Exception:
                continue
            for cell in col:
                if cell.value is not None:
                    
                    try:
                        l = len(str(cell.value))
                    except Exception:
                        l = 1
                    if l > max_len:
                        max_len = l
            sheet.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)

    
    ws_port = wb.create_sheet("PORTADA")
    ws_port.merge_cells("A1:F1")
    ws_port["A1"] = "DOCUMENTACIÓN TÉCNICA - BD STELMAN BUSES"
    ws_port["A1"].font = title_font
    ws_port["A3"] = f"Base de datos: {db_name}"
    ws_port["A3"].font = subtitle_font
    ws_port["A4"] = f"Generado el: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_port["A5"] = "Generado por: Sistema Stelman - Exportador Premium"
    ws_port["A7"] = "Contenido:"
    ws_port["A8"] = "- Índice con links a cada hoja"
    ws_port["A9"] = "- DDL completo por tabla"
    ws_port["A10"] = "- Hojas por tabla con datos, PK y FK resaltadas"
    auto_width(ws_port)

   
    ws_index = wb.create_sheet("Índice")
    headers = ["#","Tabla","Atributos","Registros","PK","FK","Estado","Link hoja"]
    ws_index.append(headers)
    for c in ws_index[1]:
        c.font = header_font
        c.fill = index_header_fill
        c.alignment = center_align
        c.border = border

    
    ws_sql = wb.create_sheet("CREATE")
    ws_sql.append(["Tabla", "CREATE TABLE (DDL)"])
    for c in ws_sql[1]:
        c.font = header_font
        c.fill = index_header_fill
        c.border = border

    
    cursor.execute("SHOW TABLES")
    rows = cursor.fetchall()
    if not rows:
        cursor.close()
        flash("❌ No se encontraron tablas en la base de datos.", "danger")
        return redirect(url_for('panel_reportes'))
    key = list(rows[0].keys())[0]
    tablas = [r[key] for r in rows]


    relations = []


    idx = 2
    for t in tablas:
        
        cursor.execute(f"SHOW CREATE TABLE `{t}`")
        ddl_row = cursor.fetchone()
        ddl = ddl_row.get("Create Table") if ddl_row and "Create Table" in ddl_row else str(ddl_row)

        
        cursor.execute("""
            SELECT COLUMN_NAME, COLUMN_KEY, COLUMN_TYPE, IS_NULLABLE, COLUMN_DEFAULT
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_SCHEMA=%s AND TABLE_NAME=%s
            ORDER BY ORDINAL_POSITION
        """, (db_name, t))
        cols = cursor.fetchall()

      
        try:
            cursor.execute(f"SELECT COUNT(*) AS c FROM `{t}`")
            cnt = cursor.fetchone()["c"]
        except Exception:
            cnt = 0

      
        pk = [c["COLUMN_NAME"] for c in cols if c["COLUMN_KEY"] == "PRI"]
        pk_str = ", ".join(pk) if pk else "—"

      
        cursor.execute("""
            SELECT COLUMN_NAME, REFERENCED_TABLE_NAME, REFERENCED_COLUMN_NAME
            FROM INFORMATION_SCHEMA.KEY_COLUMN_USAGE
            WHERE TABLE_SCHEMA=%s AND TABLE_NAME=%s
              AND REFERENCED_TABLE_NAME IS NOT NULL
        """, (db_name, t))
        fks = cursor.fetchall()
        fk_str = ", ".join([f"{f['COLUMN_NAME']} → {f['REFERENCED_TABLE_NAME']}.{f['REFERENCED_COLUMN_NAME']}" for f in fks]) or "—"

    
        estado = "OK"
        if not pk:
            estado = "⚠ Sin PK"

        ws_index.append([idx-1, t, len(cols), cnt, pk_str, fk_str, estado, f"=HYPERLINK(\"#{t}!A1\",\"Abrir\")"])
        idx += 1

   
        ws_sql.append([t, ddl])

    
        sheet_name = t[:31]
        ws = wb.create_sheet(sheet_name)

  
        ws["A1"] = f"CREATE TABLE: {t}"
        ws["A1"].font = header_font
        ws["A1"].fill = ddl_fill
        ws.merge_cells("A1:F1")

       
        ws["A2"] = ddl
        ws["A2"].alignment = Alignment(wrap_text=True)
        ws["A2"].fill = ddl_fill
        ws.merge_cells("A2:F4")
        ws.row_dimensions[2].height = 70

       
        ws["A6"] = "PK:"
        ws["B6"] = pk_str
        ws["A7"] = "FK:"
        ws["B7"] = fk_str

       
        for fk in fks:
            relations.append((t, fk["COLUMN_NAME"], fk["REFERENCED_TABLE_NAME"], fk["REFERENCED_COLUMN_NAME"]))

      
        cursor.execute(f"SELECT * FROM `{t}` LIMIT 1000")
        datos = cursor.fetchall()
        headers_data = [c["COLUMN_NAME"] for c in cols]

        row_start = 9
        col_idx = 1
        for h in headers_data:
            cell = ws.cell(row=row_start, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = index_header_fill
            cell.border = border
            cell.alignment = center_align
            col_idx += 1

  
        fila = row_start + 1
        for row in datos:
            col_idx = 1
            for h in headers_data:
                v = row.get(h)
                cell = ws.cell(row=fila, column=col_idx, value=v)
               
                if h in pk:
                    cell.fill = pk_fill
                else:
                    
                    fk_cols = [f['COLUMN_NAME'] for f in fks]
                    if h in fk_cols:
                        cell.fill = fk_fill
                cell.border = border
                col_idx += 1
            fila += 1

       
        last_col = get_column_letter(len(headers_data))
        if datos:
            ws.auto_filter.ref = f"A{row_start}:{last_col}{fila-1}"

   
        ws.freeze_panes = f"A{row_start+1}"

        
        auto_width(ws)

 
    ws_erd = wb.create_sheet("ERD_TEXTUAL")
    ws_erd["A1"] = "Mini ERD (texto)"
    ws_erd["A1"].font = header_font
    r = 3
    if relations:
        for (src_table, src_col, dst_table, dst_col) in relations:
            ws_erd.cell(row=r, column=1, value=f"{src_table}.{src_col}  →  {dst_table}.{dst_col}")
            r += 1
    else:
        ws_erd["A3"] = "No se detectaron relaciones (FK) en el export."
    auto_width(ws_erd)

   
    auto_width(ws_index)
    auto_width(ws_sql)

    
    for i, row in enumerate(ws_index.iter_rows(min_row=2, max_col=8), start=2):
        estado_cell = row[6]  
        try:
            if "Sin PK" in str(estado_cell.value):
                for c in row:
                    c.fill = PatternFill("solid", fgColor="FFD7D7")  
        except Exception:
            pass

    
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    file_name = f"DOCUMENTACION_BD_{db_name}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    cursor.close()
    return send_file(output,
                     download_name=file_name,
                     as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



#=====================================================
# EJECUCIÓN DEL SERVIDOR
#=====================================================
if __name__ == '__main__':
    app.run(debug=True)
